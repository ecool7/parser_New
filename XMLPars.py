import tkinter as tk
from tkinter import filedialog, messagebox
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import json

# Метод проверки
def get_text_or_none(element):
    if element is not None:
        text = element.text.strip() if element.text else None
        return text if text else None
    return None

def process_files(xml_file, output_file,toexel,type_file,spravka):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    data = []

    for record in root.findall('.//fields/field'):
        code = get_text_or_none(record.find('code'))
        indexNumber = get_text_or_none(record.find('indexNumber'))


        wb = load_workbook(type_file)
        sheet_ranges = wb['Лист1']
        b = {}
        for row in sheet_ranges.iter_rows(min_row=2):
            a = row[0].value
            c = row[1].value
            b[a] = c
        type_t = record.findall('type')
        type_t = [get_text_or_none(type_t) for type_t in record.findall('type')]
        type_t =  [b.get(item, item) for item in type_t]
        type_t = '\n'.join(type_t)

        # type_t = get_text_or_none(record.find('type'))

        sub_type = record.findall('subtype')
        sub_type = [get_text_or_none(sub_type) for sub_type in record.findall('subtype')]
        sub_type =  [b.get(item, item) for item in sub_type]
        sub_type = '\n'.join(sub_type)

        # sub_type = get_text_or_none(record.find('subtype'))





        title = get_text_or_none(record.find('title'))
        hidden = get_text_or_none(record.find('hidden'))
        readonly = get_text_or_none(record.find('readonly'))
        editable = get_text_or_none(record.find('editable'))
        property = get_text_or_none(record.find('master/property'))
        action = get_text_or_none(record.find('master/action'))

        code_list = [get_text_or_none(code) for code in record.findall('master/code')]
        code3 = '\n'.join(code_list) if code_list else ''
        if code3 == 'None(None.None->None)':
            code3 = ''

        if code3 == 'composeValue($sessionUser.copy->value)':
            code3 = 'sessionUser'

        var_elements = record.findall('master/var')
        var = chr(10).join([get_text_or_none(v) for v in var_elements if get_text_or_none(v) is not None])

        dependscode_elements = record.findall('dependsOnFields/dependsOn/code')
        dependscode = chr(10).join([get_text_or_none(dc) for dc in dependscode_elements if get_text_or_none(dc) is not None])
        
        key1 = record.findall('hideOnTaskDefinitionKeys/key')
        key = chr(10).join([get_text_or_none(dc) for dc in key1 if get_text_or_none(dc) is not None])

        externalTable1 = record.findall('externalTable')
        externalTable = ", ".join([get_text_or_none(dc) for dc in externalTable1 if get_text_or_none(dc) is not None])

        externalTableQuery1 = record.findall('externalTableQuery')
        externalTableQuery = ", ".join([get_text_or_none(dc) for dc in externalTableQuery1 if get_text_or_none(dc) is not None])



        # Доработки 10/06: Добавление отсылки сабформы, правило валидации(пока через запятую перечисление )
        parent_code = get_text_or_none(record.find('parentCode'))
        
        if parent_code is None:
            parent_code = ""
        else: parent_code = parent_code
        


        wb = load_workbook(toexel)
        sheet_ranges = wb['Лист1']
        b = {}
        for row in sheet_ranges.iter_rows(min_row=2):
            a = row[0].value
            c = row[1].value
            b[a] = c
        rules = record.findall('rules/rule')
        rules = [get_text_or_none(rules) for rules in record.findall('rules/rule')]
        rules =  [b.get(item, item) for item in rules]
        rule = '\n'.join(rules)



        wb = load_workbook(spravka)
        sheet_ranges = wb['Лист1']
        b = {}
        for row in sheet_ranges.iter_rows(min_row=2):
            a = row[0].value
            c = row[1].value
            b[a] = c

        dictionaryCode = record.findall('dictionaryCode')
        dictionaryCode = [get_text_or_none(dictionaryCode) for dictionaryCode in record.findall('dictionaryCode')]
        dictionaryCode =  [b.get(item_s, item_s) for item_s in dictionaryCode]
        dictionaryCode = '\n'.join(dictionaryCode)

        # dictionaryCode = get_text_or_none(record.find('dictionaryCode'))
        if dictionaryCode is None:
            dictionaryCode = ""
        else: dictionaryCode = dictionaryCode



        sourceDictionary = [get_text_or_none(sD) for sD in record.findall('externalTableFields/field/source')]
        sourceDictionary = '\n'.join(sourceDictionary) if sourceDictionary else ''

        item = record.findall('items/item/name')
        item = chr(10).join([get_text_or_none(it) for it in item if get_text_or_none(it) is not None])

        defaultValue  =  get_text_or_none(record.find('defaultValue'))
        
        if defaultValue is None:
            defaultValue = ' '
    
        else:
            if defaultValue.startswith('{'):
                defaultValue = json.loads(defaultValue)
                defaultValue = defaultValue['name']
            else:
                defaultValue = defaultValue

        


        
        data.append({   #'Номер': indexNumber, 
                    'Название полей': title, 
                    'Системное Название Поля': code,  
                    'Значение по умолчанию': defaultValue,
                    'Основной тип поля': type_t,
                    'Дополнительный тип поля': sub_type,
                    'Скрытые поля': hidden, 
                    'Только для чтения': readonly,
                    'Условия отображения': dependscode,
                    'Правила валидации': rule,
                    'Скрывать поле на этапе': key,
                    'Справочник значений': dictionaryCode,
                    'Значения справочника' : item,
                    'В какой таблице ищем': externalTable,
                    'Запрос для поиска в таблице': externalTableQuery,
                    'Сабформа' : parent_code,
                    'Мастер поля': code3

                    # '??????':sourceDictionary не используется
                    })



    df1 = pd.DataFrame(data)
    
 

    # Запись датафрейма в эксель 
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df1.to_excel(writer, index=False)
    writer.close()

    # Открытие файла для настройки форматирования
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    column_widths = {'A': 30, 'B': 30,'C': 30,'D': 30,'E': 30, 'F': 30,'G': 30,'H':30,'I':30,'J':30,'K':30,'L':30,'M':30,'N':30,'O':30,'P':30}
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Настройка форматирования для переноса строк
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
       for cell in row:
           cell.alignment = Alignment(wrap_text=True)


    # Настройка форматирования для переноса строк
    columns_to_wrap = ['H', 'I', 'J']  # Замените на ваши столбцы
    for col in columns_to_wrap:
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True)        

    # Сохранение изменений
    wb.save(output_file)


def select_xml_file():
    xml_file = filedialog.askopenfilename(title="Выберите XML файл", filetypes=[("XML Files", "*.xml")])
    if xml_file:
        xml_file_entry.delete(0, tk.END)
        xml_file_entry.insert(0, xml_file)

def select_rule_file():
    toexel = filedialog.askopenfilename(title="Выберите  файл валидации",defaultextension=".xlsx", filetypes=[("Excel files", ".xlsx .xls")])
    if toexel:
        rule_file.delete(0, tk.END)
        rule_file.insert(0, toexel)


def select_type():
    type_file = filedialog.askopenfilename(title="Выберите  файл типов",defaultextension=".xlsx", filetypes=[("Excel files", ".xlsx .xls")])
    if type_file:
        typeFL.delete(0, tk.END)
        typeFL.insert(0, type_file)

def znach():
    spravka = filedialog.askopenfilename(title="Выберите  файл типов",defaultextension=".xlsx", filetypes=[("Excel files", ".xlsx .xls")])
    if spravka:
        spravka_file.delete(0, tk.END)
        spravka_file.insert(0, spravka)

def select_output_file():
    output_file = filedialog.asksaveasfilename(title="Сохранить как", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if output_file:
        output_file_entry.delete(0, tk.END)
        output_file_entry.insert(0, output_file)

def start_processing():
    xml_file = xml_file_entry.get()
    toexel = rule_file.get()
    type_file = typeFL.get()
    output_file = output_file_entry.get()
    spravk_file = spravka_file.get()
    if not xml_file or not output_file :
        messagebox.showwarning("Ошибка", "Пожалуйста, выберите XML файл и файл для сохранения результата.")
        return
    
    
    try:
        process_files(xml_file, output_file,toexel,type_file,spravk_file)
        messagebox.showinfo("Успех", "Файл успешно обработан и сохранен.")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")
        

# Создание основного окна
root = tk.Tk()
root.title("XML to Excel Converter")
# Создание и размещение элементов интерфейса
tk.Label(root, text="Выберите XML файл:").grid(row=0, column=0, padx=10, pady=10)
xml_file_entry = tk.Entry(root, width=50)
xml_file_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Выбрать", command=select_xml_file).grid(row=0, column=2, padx=10, pady=10)




tk.Label(root, text="Выберите файл для валидации :").grid(row=1, column=0, padx=10, pady=10)
rule_file = tk.Entry(root, width=50)
rule_file.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Выбрать", command=select_rule_file).grid(row=1, column=2, padx=10, pady=10)


tk.Label(root, text="Выберите файл для типов :").grid(row=2, column=0, padx=10, pady=10)
typeFL = tk.Entry(root, width=50)
typeFL.grid(row=2, column=1, padx=10, pady=10)
tk.Button(root, text="Выбрать", command=select_type).grid(row=2, column=2, padx=10, pady=10)


tk.Label(root, text="Выберите файл для значений справочника :").grid(row=3, column=0, padx=10, pady=10)
spravka_file = tk.Entry(root, width=50)
spravka_file.grid(row=3, column=1, padx=10, pady=10)
tk.Button(root, text="Выбрать", command=znach).grid(row=3, column=2, padx=10, pady=10)


tk.Label(root, text="Сохранить результат как:").grid(row=4, column=0, padx=10, pady=10)
output_file_entry = tk.Entry(root, width=50)
output_file_entry.grid(row=4, column=1, padx=10, pady=10)
tk.Button(root, text="Выбрать", command=select_output_file).grid(row=4, column=2, padx=10, pady=10)






tk.Button(root, text="Начать обработку", command=start_processing).grid(row=5, column=0, columnspan=3, pady=20)

root.mainloop()
